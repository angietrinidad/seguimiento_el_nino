"""
Excel de SERVICIOS DE SALUD en riesgo de inundación (fluvial), análogo al de
escuelas. Fuente: MSPBS (categoría/subtipo/distrito) cruzado con las zonas de
inundación derivadas de Sentinel-1.

Salida: datos/exposicion/salud_riesgo.xlsx
"""
import os, json
from shapely.geometry import shape, Point
from shapely.ops import unary_union
from shapely.prepared import prep
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GEO = os.path.join(os.path.dirname(__file__), "..", "datos", "geo")
EXP = os.path.join(os.path.dirname(__file__), "..", "datos", "exposicion")

def huella(fn):
    gs = [shape(f["geometry"]).buffer(0) for f in
          json.load(open(os.path.join(GEO, fn), encoding="utf-8"))["features"]]
    return prep(unary_union(gs).buffer(0.005))
h2016 = huella("inundacion-2015-16-s1.geojson")
h2019 = huella("inundacion-2018-19-s1.geojson")

def nivel_rio(zona, pt):
    if zona.startswith("Asunción"):
        if h2019.contains(pt): return "≈ 7,57 m (se inundó en 2018-19)"
        if h2016.contains(pt): return "≈ 7,88 m (se inundó en 2015-16)"
        return "≥ 7,5 m (zona de riesgo)"
    return "zona de anegamiento recurrente (estación local)"

exp = json.load(open(os.path.join(GEO, "expuestos-inundacion.geojson"), encoding="utf-8"))["features"]
filas = []
for f in exp:
    p = f["properties"]
    if p.get("clase") != "salud":
        continue
    lon, lat = f["geometry"]["coordinates"]
    filas.append({
        "nombre": p.get("nombre", ""),
        "categoria": p.get("categoria", ""), "subtipo": p.get("subtipo", ""),
        "distrito": p.get("distrito", ""), "lat": round(lat, 5), "lon": round(lon, 5),
        "zona": p.get("zona", ""), "nivel_rio": nivel_rio(p.get("zona", ""), Point(lon, lat))})
filas.sort(key=lambda r: (r["zona"], r["categoria"], r["nombre"]))
json.dump(filas, open(os.path.join(EXP, "salud_riesgo.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)

# ---- Excel ----
wb = Workbook(); ws = wb.active; ws.title = "Salud en riesgo"
COLS = [("Servicio de salud", "nombre", 40), ("Categoría", "categoria", 30),
        ("Subtipo", "subtipo", 12), ("Distrito", "distrito", 26),
        ("Latitud", "lat", 11), ("Longitud", "lon", 11),
        ("Zona de riesgo", "zona", 26), ("Se inunda con nivel del río", "nivel_rio", 34)]
azul = PatternFill("solid", fgColor="1F4E79"); blanco = Font(color="FFFFFF", bold=True)
borde = Border(*[Side(style="thin", color="D9D9D9")] * 4); zebra = PatternFill("solid", fgColor="F2F6FB")

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
c = ws.cell(1, 1, "Servicios de salud (MSPBS) en zonas de riesgo de inundación — Paraguay")
c.font = Font(bold=True, size=13, color="1F4E79")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
c = ws.cell(2, 1, ("Fuente: MSPBS (establecimientos georreferenciados) cruzado con zonas de inundación "
                   "Sentinel-1. Cada servicio se cuenta en una sola zona (sin doble conteo)."))
c.font = Font(size=9, italic=True, color="595959"); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 30

HR = 4
for j, (t, _, w) in enumerate(COLS, 1):
    cell = ws.cell(HR, j, t); cell.fill = azul; cell.font = blanco; cell.border = borde
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HR].height = 28
for i, r in enumerate(filas):
    for j, (_, k, _) in enumerate(COLS, 1):
        cell = ws.cell(HR + 1 + i, j, r.get(k) or "s/d"); cell.border = borde
        cell.alignment = Alignment(horizontal="center" if k in ("lat", "lon") else "left",
                                   vertical="center", wrap_text=k in ("nombre", "nivel_rio"))
        if i % 2: cell.fill = zebra
ws.freeze_panes = f"A{HR+1}"
ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(COLS))}{HR+len(filas)}"

out = os.path.join(EXP, "salud_riesgo.xlsx"); wb.save(out)
print(f"Excel de salud guardado: {out} | {len(filas)} servicios")
from collections import Counter
print("por categoría:", dict(Counter(r["categoria"] for r in filas)))
