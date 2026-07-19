"""
Excel de ACTIVIDAD COMERCIAL en riesgo de inundación (fluvial), análogo a los de
escuelas y salud. Fuente: OpenStreetMap (shop=* + marketplace) cruzado con las zonas
de inundación derivadas de Sentinel-1.

Salida: datos/exposicion/comercio_riesgo.xlsx
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GEO = os.path.join(os.path.dirname(__file__), "..", "datos", "geo")
EXP = os.path.join(os.path.dirname(__file__), "..", "datos", "exposicion")
feats = json.load(open(os.path.join(GEO, "comercio-expuesto.geojson"), encoding="utf-8"))["features"]
filas = [{"nombre": f["properties"].get("nombre") or "(sin nombre)",
          "tipo": f["properties"].get("tipo", ""), "shop": f["properties"].get("shop", ""),
          "lat": round(f["geometry"]["coordinates"][1], 5),
          "lon": round(f["geometry"]["coordinates"][0], 5)} for f in feats]
filas.sort(key=lambda r: (r["tipo"], r["nombre"]))

wb = Workbook(); ws = wb.active; ws.title = "Comercio en riesgo"
COLS = [("Comercio", "nombre", 40), ("Tipo", "tipo", 26), ("Rubro OSM (shop)", "shop", 20),
        ("Latitud", "lat", 11), ("Longitud", "lon", 11)]
azul = PatternFill("solid", fgColor="1F4E79"); blanco = Font(color="FFFFFF", bold=True)
borde = Border(*[Side(style="thin", color="D9D9D9")] * 4); zebra = PatternFill("solid", fgColor="F2F6FB")

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
c = ws.cell(1, 1, "Actividad comercial (OSM) en la zona de riesgo de inundación — Asunción metro")
c.font = Font(bold=True, size=13, color="1F4E79")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
c = ws.cell(2, 1, ("Fuente: OpenStreetMap (comercios y mercados) cruzado con zonas de inundación "
                   "Sentinel-1. Cota INFERIOR: OSM subrepresenta el comercio informal y los barrios sin mapear."))
c.font = Font(size=9, italic=True, color="595959"); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 30

HR = 4
for j, (t, _, w) in enumerate(COLS, 1):
    cell = ws.cell(HR, j, t); cell.fill = azul; cell.font = blanco; cell.border = borde
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = w
for i, r in enumerate(filas):
    for j, (_, k, _) in enumerate(COLS, 1):
        cell = ws.cell(HR + 1 + i, j, r.get(k) or "s/d"); cell.border = borde
        cell.alignment = Alignment(horizontal="center" if k in ("lat", "lon") else "left", vertical="center")
        if i % 2: cell.fill = zebra
ws.freeze_panes = f"A{HR+1}"
ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(COLS))}{HR+len(filas)}"
out = os.path.join(EXP, "comercio_riesgo.xlsx"); wb.save(out)
print(f"Excel de comercio guardado: {out} | {len(filas)} comercios")
