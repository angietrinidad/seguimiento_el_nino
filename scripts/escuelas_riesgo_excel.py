"""
Construye el Excel de ESCUELAS EN RIESGO de inundación a partir de
datos/exposicion/escuelas_riesgo.json (generado por escuelas_riesgo_tabla.py).

Salida: datos/exposicion/escuelas_riesgo.xlsx
Columnas: institución, departamento, distrito, barrio, lat, lon,
matrícula (último año), año, tipo de gestión, zona de riesgo, nivel del río.
"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXP = os.path.join(os.path.dirname(__file__), "..", "datos", "exposicion")
filas = json.load(open(os.path.join(EXP, "escuelas_riesgo.json"), encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "Escuelas en riesgo"

COLS = [
    ("Institución educativa", "institucion", 42),
    ("Departamento", "departamento", 16),
    ("Distrito", "distrito", 18),
    ("Barrio / localidad", "barrio", 22),
    ("Latitud", "lat", 11),
    ("Longitud", "lon", 11),
    ("Matrícula", "matricula", 11),
    ("Año matrícula", "anio_matricula", 13),
    ("Tipo de gestión", "gestion", 22),
    ("Zona de riesgo", "zona", 26),
    ("Se inunda con nivel del río", "nivel_rio", 34),
]

azul = PatternFill("solid", fgColor="1F4E79")
blanco = Font(color="FFFFFF", bold=True, size=11)
borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)
zebra = PatternFill("solid", fgColor="F2F6FB")

# Título
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
c = ws.cell(1, 1, "Establecimientos educativos en zonas de riesgo de inundación — Paraguay")
c.font = Font(bold=True, size=13, color="1F4E79")
c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 24

sub = ("Fuente: MEC (establecimientos y matrícula) cruzado con zonas de inundación derivadas de Sentinel-1. "
       "Cada escuela se cuenta en una sola zona (sin doble conteo en solapamientos). "
       "Nivel del río: observado por episodio en Asunción; en el resto, zona de anegamiento recurrente por estación local.")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
c = ws.cell(2, 1, sub)
c.font = Font(size=9, italic=True, color="595959")
c.alignment = Alignment(vertical="top", wrap_text=True)
ws.row_dimensions[2].height = 42

# Encabezados
HR = 4
for j, (titulo, _, ancho) in enumerate(COLS, 1):
    cell = ws.cell(HR, j, titulo)
    cell.fill = azul; cell.font = blanco
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borde
    ws.column_dimensions[get_column_letter(j)].width = ancho
ws.row_dimensions[HR].height = 30

# Datos
for i, r in enumerate(filas):
    fila = HR + 1 + i
    for j, (_, key, _) in enumerate(COLS, 1):
        v = r.get(key)
        if v is None:
            v = "s/d"
        cell = ws.cell(fila, j, v)
        cell.border = borde
        cell.alignment = Alignment(
            horizontal="center" if key in ("lat", "lon", "matricula", "anio_matricula") else "left",
            vertical="center", wrap_text=key in ("institucion", "nivel_rio"))
        if i % 2:
            cell.fill = zebra
    ws.row_dimensions[fila].height = 28

# Fila de total
tot_fila = HR + 1 + len(filas)
total_mat = sum(r["matricula"] or 0 for r in filas)
ws.cell(tot_fila, 1, f"TOTAL — {len(filas)} establecimientos").font = Font(bold=True)
c = ws.cell(tot_fila, 7, total_mat); c.font = Font(bold=True)
c.alignment = Alignment(horizontal="center")
for j in range(1, len(COLS) + 1):
    ws.cell(tot_fila, j).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(tot_fila, j).border = borde

ws.freeze_panes = f"A{HR+1}"
ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(COLS))}{tot_fila-1}"

# Hoja 2: notas metodológicas
ws2 = wb.create_sheet("Notas")
notas = [
    ("Metodología y advertencias", ""),
    ("", ""),
    ("Establecimientos", "Establecimientos educativos oficiales del MEC (portal datos.mec.gov.py) cuyas coordenadas caen dentro de una zona de riesgo de inundación."),
    ("Zonas de riesgo", "Unión de huellas de inundación detectadas con radar Sentinel-1 (episodios 2015-16, 2018-19, 2023-24 en Asunción; anegamiento recurrente en Villa Hayes, Alberdi, Ñeembucú) + buffer de ~550 m."),
    ("Sin doble conteo", "Villa Hayes y el área metropolitana de Asunción se solapan geográficamente. Cada establecimiento se asigna a una única zona (la más local/específica primero), por lo que la suma no repite escuelas."),
    ("Matrícula", "Cantidad de matriculados del último año disponible en el MEC para las instituciones que funcionan en el establecimiento. 's/d' = sin dato en la fuente."),
    ("Tipo de gestión", "Inferido del nombre oficial de la institución: 'PRIV.SUBV.' = privada subvencionada; 'PRIV.' = privada; caso contrario = oficial (pública)."),
    ("Nivel del río", "En Asunción: nivel observado del río Paraguay en el episodio cuya huella cubre la escuela (≈7,57 m en 2018-19; ≈7,88 m en 2015-16; etapa crítica oficial: 5,5 m). En el resto: 'zona de anegamiento recurrente' según la estación hidrométrica local (Villa Hayes, Alberdi, Pilar)."),
    ("Advertencia", "Cifras trazables a fuente primaria (MEC + Sentinel-1). No sustituyen un relevamiento en terreno. Los datos sin fuente verificable se marcan 's/d' y no se estiman."),
]
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 95
for i, (a, b) in enumerate(notas, 1):
    ca = ws2.cell(i, 1, a); cb = ws2.cell(i, 2, b)
    if i == 1:
        ca.font = Font(bold=True, size=13, color="1F4E79")
    else:
        ca.font = Font(bold=True)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[i].height = 48 if len(b) > 90 else 18

out = os.path.join(EXP, "escuelas_riesgo.xlsx")
wb.save(out)
print("Excel guardado:", out)
print(f"{len(filas)} escuelas, matrícula total {total_mat:,}")
